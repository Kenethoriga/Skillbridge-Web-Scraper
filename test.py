
"""
SkillBridge Data Extractor - Filter-Triggered Version
Interacts with dropdowns to load data, then extracts
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import pandas as pd
from datetime import datetime
import time
import os
import re
from openpyxl.styles import Font, Alignment, PatternFill

class SkillBridgeExtractor:
    def __init__(self, headless=False):
        self.headless = headless
        self.driver = None
        self.wait = None
        self.base_url = "https://skillbridge.osd.mil/locations.htm"
        self.setup_driver()
    
    def setup_driver(self):
        """Setup Chrome driver"""
        print("🛠️  Setting up Chrome driver...")
        chrome_options = Options()
        
        if self.headless:
            chrome_options.add_argument("--headless=new")
        
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Additional options for better stability
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-plugins")
        chrome_options.add_argument("--disable-images")
        chrome_options.add_argument("--disable-javascript")
        
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            self.wait = WebDriverWait(self.driver, 60)
            print("✅ Chrome driver setup complete")
        except Exception as e:
            print(f"❌ Failed to setup Chrome driver: {e}")
            raise
    
    def navigate_to_page(self):
        """Navigate to SkillBridge locations page"""
        print(f"\n🌐 Navigating to {self.base_url}...")
        try:
            self.driver.get(self.base_url)
            time.sleep(5)  # Wait for initial load
            
            # Handle potential popups
            self.handle_cookies_popup()
            self.handle_popups()
            
            print("✅ Page loaded")
            return True
        except Exception as e:
            print(f"❌ Error navigating to page: {e}")
            return False
    
    def handle_cookies_popup(self):
        """Handle any cookie consent popups"""
        try:
            cookie_selectors = [
                "button[aria-label*='cookie']",
                "button[class*='cookie']",
                "#accept-cookies",
                ".cookie-accept",
                "button:contains('Accept')",
                "#gdpr-accept",
                ".accept-cookies"
            ]
            for selector in cookie_selectors:
                try:
                    button = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
                    if button.is_displayed():
                        button.click()
                        print("✅ Handled cookie popup")
                        time.sleep(1)
                        return True
                except:
                    continue
            return False
        except Exception as e:
            print(f"⚠️  No cookie popup or error handling it: {e}")
            return False
    
    def handle_popups(self):
        """Handle other potential popups"""
        try:
            popup_selectors = [
                ".modal-close",
                ".close-button",
                "[aria-label='Close']",
                ".btn-close",
                "#close-modal"
            ]
            for selector in popup_selectors:
                try:
                    button = self.driver.find_element(By.CSS_SELECTOR, selector)
                    if button.is_displayed():
                        button.click()
                        print("✅ Closed popup")
                        time.sleep(1)
                        break
                except:
                    continue
        except:
            pass
    
    def wait_for_page_ready(self):
        """Wait for page to be fully loaded"""
        print("⏳ Waiting for page to be ready...")
        
        max_wait = 30
        start_time = time.time()
        
        while time.time() - start_time < max_wait:
            try:
                # Wait for jQuery and DataTables
                ready = self.driver.execute_script(
                    "return typeof jQuery !== 'undefined' && jQuery.active === 0"
                )
                
                # Wait for the industries dropdown (Job Family)
                dropdown_present = EC.presence_of_element_located((By.ID, "industries-dropdown"))
                WebDriverWait(self.driver, 5).until(dropdown_present)
                
                # Check if DataTables is initialized
                dt_ready = self.driver.execute_script(
                    "return typeof jQuery.fn.DataTable !== 'undefined'"
                )
                
                if ready and dt_ready:
                    print("✅ Page ready")
                    time.sleep(2)
                    return True
                    
            except Exception as e:
                print(f"⏳ Still loading... ({int(time.time() - start_time)}s)")
            
            time.sleep(1)
        
        print("⚠️  Page ready timeout, continuing anyway...")
        return False
    
    def get_industry_options(self):
        """Get all available options from industries dropdown"""
        print("\n🔍 Checking available Job Family options...")
        
        try:
            dropdown = Select(self.driver.find_element(By.ID, "industries-dropdown"))
            options = dropdown.options
            
            # Get option texts (skip first if it's "All" or empty)
            available_options = []
            for option in options:
                text = option.text.strip()
                if text and text.lower() not in ['all', 'select', '']:
                    available_options.append(text)
            
            print(f"✅ Found {len(available_options)} job family options:")
            for opt in available_options:
                print(f"   - {opt}")
            
            return available_options
            
        except Exception as e:
            print(f"❌ Error getting dropdown options: {e}")
            return []
    
    def select_industry(self, industry_name):
        """Select a specific industry from dropdown"""
        print(f"\n🎯 Selecting industry: {industry_name}")
        
        try:
            # Find the dropdown
            dropdown = Select(self.driver.find_element(By.ID, "industries-dropdown"))
            
            # Try exact match first
            try:
                dropdown.select_by_visible_text(industry_name)
                print(f"✅ Selected: {industry_name}")
                time.sleep(2)  # Wait for selection to apply
                return True
            except:
                # Try partial match
                for option in dropdown.options:
                    if industry_name.lower() in option.text.lower():
                        actual_text = option.text
                        dropdown.select_by_visible_text(actual_text)
                        print(f"✅ Selected (partial match): {actual_text}")
                        time.sleep(2)
                        return True
                
                print(f"⚠️  Could not find '{industry_name}' in dropdown")
                return False
                
        except Exception as e:
            print(f"❌ Error selecting industry: {e}")
            return False
    
    def trigger_search(self):
        """Trigger the search to load data"""
        print("🔍 Triggering search...")
        
        try:
            # Look for search/submit button
            button_selectors = [
                "#search-button",
                "button[type='submit']",
                "button.search",
                "input[type='submit']",
                ".btn-search",
                ".btn-primary",
                "button:contains('Search')",
                "button:contains('Filter')",
                "button:contains('Apply')"
            ]
            
            for selector in button_selectors:
                try:
                    button = self.driver.find_element(By.CSS_SELECTOR, selector)
                    if button.is_displayed() and button.is_enabled():
                        self.driver.execute_script("arguments[0].click();", button)
                        print(f"✅ Clicked search button: {selector}")
                        time.sleep(3)
                        return True
                except:
                    continue
            
            # If no button found, try JavaScript trigger
            try:
                self.driver.execute_script("""
                    jQuery('#industries-dropdown').trigger('change');
                    if (typeof jQuery.fn.DataTable !== 'undefined') {
                        jQuery('#location-table').DataTable().draw();
                    }
                """)
                print("✅ Triggered search via JavaScript")
                time.sleep(3)
                return True
            except:
                pass
            
            # If no button found, the table might auto-update on dropdown change
            print("ℹ️  No search button found - table may auto-update")
            time.sleep(5)  # Give time for auto-update
            return True
            
        except Exception as e:
            print(f"⚠️  Error triggering search: {e}")
            return False
    
    def check_for_errors(self):
        """Check for error messages or no results"""
        try:
            error_selectors = [
                ".error", ".no-results", ".dataTables_empty",
                "td:contains('No data')", "div:contains('No results')",
                "div:contains('No matching')", "div:contains('error')",
                ".alert-error", ".alert-danger"
            ]
            for selector in error_selectors:
                try:
                    element = self.driver.find_element(By.CSS_SELECTOR, selector)
                    if element.is_displayed():
                        error_text = element.text.strip()
                        if error_text and len(error_text) > 0:
                            return error_text
                except:
                    continue
            return None
        except:
            return None
    
    def wait_for_data_load(self):
        """Wait for data to load in table"""
        print("⏳ Waiting for data to load...")
        
        max_wait = 30
        start_time = time.time()
        
        while time.time() - start_time < max_wait:
            try:
                # Check for errors first
                error = self.check_for_errors()
                if error:
                    print(f"⚠️  Error message detected: {error}")
                    return 0
                
                # Check if table has data rows using multiple methods
                
                # Method 1: Check DataTables API
                try:
                    row_count = self.driver.execute_script("""
                        if (typeof jQuery.fn.DataTable !== 'undefined') {
                            var table = jQuery('#location-table').DataTable();
                            return table.rows({search: 'applied'}).count();
                        }
                        return 0;
                    """)
                    if row_count > 0:
                        print(f"✅ Data loaded - {row_count} rows via DataTables")
                        time.sleep(2)
                        return row_count
                except:
                    pass
                
                # Method 2: Check HTML table
                table = self.driver.find_element(By.ID, "location-table")
                tbody = table.find_element(By.TAG_NAME, "tbody")
                rows = tbody.find_elements(By.TAG_NAME, "tr")
                
                # Filter out "no data" or loading messages
                data_rows = []
                for r in rows:
                    try:
                        if (r.is_displayed() and 
                            'dataTables_empty' not in (r.get_attribute('class') or '') and
                            r.text.strip() and 
                            'no data' not in r.text.lower() and
                            'loading' not in r.text.lower()):
                            data_rows.append(r)
                    except:
                        continue
                
                if len(data_rows) > 0:
                    print(f"✅ Data loaded - {len(data_rows)} visible rows")
                    time.sleep(2)
                    return len(data_rows)
                
                # Check if still processing
                processing = self.driver.execute_script("""
                    return jQuery && jQuery.active > 0;
                """)
                if processing:
                    print("⏳ Data still processing...")
                
            except Exception as e:
                print(f"⏳ Waiting for table... ({int(time.time() - start_time)}s)")
            
            time.sleep(1)
        
        print("⚠️  Timeout waiting for data")
        return 0
    
    def extract_table_data(self, job_family_name):
        """Extract all visible data from table"""
        print(f"\n📊 Extracting data for {job_family_name}...")
        
        try:
            # Method 1: Use DataTables API (most reliable)
            print("   Trying DataTables API extraction...")
            programs = self.extract_via_datatables(job_family_name)
            if programs:
                return programs
            
            # Method 2: Fallback to HTML extraction
            print("   Falling back to HTML extraction...")
            programs = self.extract_via_html(job_family_name)
            if programs:
                return programs
            
            print("❌ All extraction methods failed")
            return []
            
        except Exception as e:
            print(f"❌ Error extracting table data: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def extract_via_datatables(self, job_family_name):
        """Extract data using DataTables API"""
        try:
            extract_script = """
            if (typeof jQuery.fn.DataTable === 'undefined') {
                return {error: 'DataTables not available'};
            }
            
            var table = jQuery('#location-table').DataTable();
            var data = [];
            var headers = [];
            
            // Get headers
            table.columns().every(function() {
                var header = jQuery(this.header());
                headers.push(header.text().trim());
            });
            
            // Get all rows (including filtered)
            table.rows({page: 'all', search: 'applied'}).every(function() {
                var rowData = this.data();
                var rowNode = this.node();
                
                // Convert to plain array and clean data
                var cleanData = [];
                for (var i = 0; i < rowData.length; i++) {
                    var value = rowData[i];
                    if (typeof value === 'string') {
                        // Remove HTML tags
                        value = value.replace(/<[^>]*>/g, '').trim();
                    }
                    cleanData.push(value);
                }
                
                data.push({
                    raw_data: cleanData,
                    is_displayed: jQuery(rowNode).is(':visible')
                });
            });
            
            return {
                headers: headers,
                data: data,
                total_rows: data.length
            };
            """
            
            result = self.driver.execute_script(extract_script)
            
            if not result or 'error' in result:
                return []
            
            headers = result.get('headers', [])
            data_rows = result.get('data', [])
            
            if not data_rows:
                return []
            
            programs = []
            for row_info in data_rows:
                if not row_info.get('is_displayed', True):
                    continue
                    
                row_data = row_info.get('raw_data', [])
                program = {}
                
                for i, value in enumerate(row_data):
                    header = headers[i] if i < len(headers) else f"Column_{i}"
                    program[header] = value
                
                program['Job Family'] = job_family_name
                program['Last Updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                programs.append(program)
            
            print(f"✅ Extracted {len(programs)} rows via DataTables API")
            return programs
            
        except Exception as e:
            print(f"⚠️  DataTables API extraction failed: {e}")
            return []
    
    def extract_via_html(self, job_family_name):
        """Fallback: Extract from HTML table"""
        print("   Extracting from visible HTML rows...")
        
        programs = []
        
        try:
            table = self.driver.find_element(By.ID, "location-table")
            tbody = table.find_element(By.TAG_NAME, "tbody")
            rows = tbody.find_elements(By.TAG_NAME, "tr")
            
            # Get headers
            headers = []
            try:
                thead = table.find_element(By.TAG_NAME, "thead")
                header_cells = thead.find_elements(By.TAG_NAME, "th")
                headers = [h.text.strip() for h in header_cells if h.text.strip()]
            except:
                # Estimate headers based on first data row
                if rows:
                    first_row = rows[0]
                    cells = first_row.find_elements(By.TAG_NAME, "td")
                    headers = [f"Column_{i}" for i in range(len(cells))]
            
            for row in rows:
                try:
                    if not row.is_displayed():
                        continue
                    
                    row_class = row.get_attribute('class') or ''
                    if 'dataTables_empty' in row_class:
                        continue
                    
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if not cells:
                        continue
                    
                    program = {}
                    for i, cell in enumerate(cells):
                        header = headers[i] if i < len(headers) else f"Column_{i}"
                        
                        # Get text content
                        text = cell.text.strip()
                        
                        # Try to extract links
                        try:
                            links = cell.find_elements(By.TAG_NAME, "a")
                            for link_idx, link in enumerate(links):
                                href = link.get_attribute('href')
                                if href:
                                    if len(links) == 1:
                                        program[f"{header}_Link"] = href
                                    else:
                                        program[f"{header}_Link_{link_idx+1}"] = href
                        except:
                            pass
                        
                        program[header] = text
                    
                    program['Job Family'] = job_family_name
                    program['Last Updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    programs.append(program)
                    
                except Exception as e:
                    continue
            
            print(f"✅ Extracted {len(programs)} programs from HTML")
            return programs
            
        except Exception as e:
            print(f"❌ HTML extraction error: {e}")
            return []
    
    def extract_with_retry(self, job_family_name, job_family_option, max_retries=3):
        """Extract data with retry logic"""
        for attempt in range(max_retries):
            print(f"\n🔄 Attempt {attempt + 1} of {max_retries}")
            
            try:
                programs = self.extract_job_family(job_family_name, job_family_option)
                if programs:
                    return programs
                
                print(f"⚠️  Attempt {attempt + 1} returned no data, retrying...")
                if attempt < max_retries - 1:
                    time.sleep(3)
                    
            except Exception as e:
                print(f"⚠️  Attempt {attempt + 1} failed: {e}")
                if attempt < max_retries - 1:
                    time.sleep(3)
        
        print(f"❌ All {max_retries} attempts failed for {job_family_name}")
        return []
    
    def extract_job_family(self, job_family_name, job_family_option):
        """Extract data for one job family"""
        print(f"\n{'='*70}")
        print(f"  EXTRACTING: {job_family_name}")
        print(f"{'='*70}")
        
        try:
            # Reset page
            print("🔄 Refreshing page...")
            self.driver.refresh()
            time.sleep(5)
            self.wait_for_page_ready()
            
            # Handle popups again after refresh
            self.handle_cookies_popup()
            self.handle_popups()
            
            # Select the industry
            if not self.select_industry(job_family_option):
                return []
            
            # Trigger search
            self.trigger_search()
            
            # Wait for data to load
            row_count = self.wait_for_data_load()
            
            if row_count == 0:
                print(f"⚠️  No data loaded for {job_family_name}")
                return []
            
            # Take screenshot for verification
            try:
                screenshot_file = f"screenshot_{job_family_name.replace(' ', '_').replace('/', '_')}.png"
                self.driver.save_screenshot(screenshot_file)
                print(f"📸 Screenshot saved: {screenshot_file}")
            except:
                print("⚠️  Could not save screenshot")
            
            # Extract data
            programs = self.extract_table_data(job_family_name)
            
            return programs
            
        except Exception as e:
            print(f"❌ Error extracting {job_family_name}: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def consolidate_locations(self, df):
        """Consolidate multiple locations for same organization/program"""
        if df.empty:
            return df
        
        print(f"\n🔄 Consolidating locations...")
        print(f"   Before: {len(df)} rows")
        
        # Find key columns (avoid duplicates)
        org_cols = []
        prog_cols = []
        loc_cols = []
        
        for col in df.columns:
            col_lower = col.lower()
            if any(term in col_lower for term in ['partner', 'organization', 'employer', 'company']) and col not in org_cols:
                org_cols.append(col)
            elif any(term in col_lower for term in ['program', 'opportunity']) and col not in prog_cols:
                prog_cols.append(col)
            elif any(term in col_lower for term in ['location', 'city', 'state', 'zip', 'address']) and col not in loc_cols:
                loc_cols.append(col)
        
        if not org_cols:
            print("⚠️  Could not identify organization column - skipping consolidation")
            return df
        
        # Build unique group columns
        group_cols = list(set(org_cols + prog_cols))
        print(f"   Grouping by: {', '.join(group_cols)}")
        
        # Add other non-location columns (avoid duplicates)
        for col in df.columns:
            if (col not in group_cols and col not in loc_cols and 
                col not in ['Last Updated', 'Job Family'] and '_Link' not in col):
                group_cols.append(col)
        
        # Remove any duplicates from group_cols
        group_cols = list(dict.fromkeys(group_cols))
        
        # Aggregation
        agg_dict = {}
        for col in df.columns:
            if col in loc_cols:
                agg_dict[col] = lambda x: ' | '.join(sorted(set(str(v).strip() 
                                                                 for v in x if pd.notna(v) and str(v).strip())))
            elif col == 'Last Updated':
                agg_dict[col] = 'max'
            elif col == 'Job Family':
                agg_dict[col] = 'first'
            elif col not in group_cols:
                agg_dict[col] = 'first'
        
        try:
            consolidated = df.groupby(group_cols, as_index=False).agg(agg_dict)
            print(f"   After: {len(consolidated)} rows")
            print(f"✅ Consolidated {len(df) - len(consolidated)} duplicates")
            return consolidated
        except Exception as e:
            print(f"⚠️  Consolidation error: {e}")
            print(f"   Skipping consolidation, returning original data")
            return df
    
    def save_to_excel(self, df, filename, sheet_name):
        """Save to Excel with formatting"""
        if df.empty:
            print(f"⚠️  No data to save")
            return
        
        print(f"\n💾 Saving to {filename}...")
        
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                
                worksheet = writer.sheets[sheet_name[:31]]
                
                # Header formatting
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Auto-adjust columns
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 60)
                    worksheet.column_dimensions[column_letter].width = max(adjusted_width, 12)
                
                # Text wrapping
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                worksheet.freeze_panes = 'A2'
            
            print(f"✅ Saved {len(df)} programs to {filename}")
            return True
            
        except Exception as e:
            print(f"❌ Error saving to Excel: {e}")
            return False
    
    def run_extraction(self, job_family_mappings, output_dir='output'):
        """Main extraction process"""
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        all_results = {}
        
        try:
            # Navigate and prepare
            if not self.navigate_to_page():
                print("❌ Failed to navigate to page")
                return all_results
                
            self.wait_for_page_ready()
            
            # Get available options
            available_options = self.get_industry_options()
            
            # Verify our target options exist
            verified_mappings = {}
            for job_family_name, job_family_option in job_family_mappings.items():
                found = False
                for available_opt in available_options:
                    if job_family_option.lower() in available_opt.lower():
                        verified_mappings[job_family_name] = available_opt
                        found = True
                        print(f"✅ Verified: {job_family_name} -> {available_opt}")
                        break
                
                if not found:
                    print(f"⚠️  Could not find '{job_family_option}' in dropdown for {job_family_name}")
            
            if not verified_mappings:
                print("❌ No valid job families found to extract")
                return all_results
            
            # Process each job family
            for job_family_name, job_family_option in verified_mappings.items():
                programs = self.extract_with_retry(job_family_name, job_family_option)
                
                if programs:
                    # Convert to DataFrame
                    df = pd.DataFrame(programs)
                    
                    # Consolidate
                    df = self.consolidate_locations(df)
                    
                    # Save
                    safe_name = re.sub(r'[^\w\s-]', '', job_family_name).replace(' ', '_')
                    
                    # Main file
                    main_file = os.path.join(output_dir, f"SkillBridge_{safe_name}.xlsx")
                    if self.save_to_excel(df, main_file, job_family_name):
                        all_results[job_family_name] = df
                    
                    # Backup
                    backup_file = os.path.join(output_dir, f"SkillBridge_{safe_name}_{timestamp}.xlsx")
                    self.save_to_excel(df, backup_file, job_family_name)
                    
                    # CSV
                    csv_file = main_file.replace('.xlsx', '.csv')
                    df.to_csv(csv_file, index=False, encoding='utf-8')
                    print(f"✅ CSV: {csv_file}")
                else:
                    print(f"❌ No programs extracted for {job_family_name}")
            
            # Summary
            print(f"\n{'='*70}")
            print("  EXTRACTION COMPLETE!")
            print(f"{'='*70}")
            total_programs = 0
            for name, df in all_results.items():
                print(f"✅ {name}: {len(df)} programs")
                total_programs += len(df)
            print(f"\n📊 TOTAL: {total_programs} programs across {len(all_results)} categories")
            print(f"📁 Files saved to: {os.path.abspath(output_dir)}")
            print(f"{'='*70}")
            
        except Exception as e:
            print(f"\n❌ FATAL ERROR: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            if not self.headless and all_results:
                print("\n⏳ Browser will close in 10 seconds...")
                time.sleep(10)
            self.close()
        
        return all_results
    
    def close(self):
        """Close browser"""
        if self.driver:
            try:
                self.driver.quit()
                print("🔚 Browser closed")
            except:
                pass

def main():
    print("""
    ╔══════════════════════════════════════════════════════════════╗
    ║     SKILLBRIDGE DATA EXTRACTOR - DROPDOWN VERSION            ║
    ║     Extracts: Sales & Business/Financial Operations          ║
    ╚══════════════════════════════════════════════════════════════╝
    """)
    
    # Map friendly names to expected dropdown options
    job_family_mappings = {
        'Sales and Related': 'Sales and Related',
        'Business and Financial Operations': 'Business and Financial Operations'
    }
    
    print("\n📋 NOTE: The script will first show you ALL available job family options")
    print("   from the dropdown. If the exact names don't match, we'll use the")
    print("   closest available options.\n")
    
    # User choice
    print("Run Options:")
    print("1. Headless mode (background - faster)")
    print("2. Visible mode (RECOMMENDED - see it work!)")
    
    try:
        choice = input("\nSelect (1 or 2, default=2): ").strip() or '2'
        headless = choice == '1'
    except:
        headless = False
    
    # Run extraction
    extractor = SkillBridgeExtractor(headless=headless)
    
    print("\n" + "="*70)
    print("STEP 1: Checking available dropdown options...")
    print("="*70)
    
    extractor.navigate_to_page()
    extractor.wait_for_page_ready()
    available_options = extractor.get_industry_options()
    
    # Check if our target options exist
    print("\n" + "="*70)
    print("STEP 2: Verifying target job families...")
    print("="*70)
    
    # Find Business and Financial Operations with fuzzy matching
    business_match = None
    for opt in available_options:
        if 'business' in opt.lower() and 'financial' in opt.lower():
            business_match = opt
            print(f"✅ Found 'Business and Financial Operations' as: '{opt}'")
            job_family_mappings['Business and Financial Operations'] = opt
            break
    
    if not business_match:
        print("⚠️  'Business and Financial Operations' not found in dropdown")
        print("   Available options containing 'Business':")
        for opt in available_options:
            if 'business' in opt.lower():
                print(f"   - {opt}")
    
    # Check Sales
    sales_match = None
    for opt in available_options:
        if 'sales' in opt.lower():
            sales_match = opt
            print(f"✅ Found 'Sales and Related' as: '{opt}'")
            job_family_mappings['Sales and Related'] = opt
            break
    
    if not sales_match:
        print("⚠️  'Sales and Related' not found in dropdown")
        print("   Available options containing 'Sales':")
        for opt in available_options:
            if 'sales' in opt.lower():
                print(f"   - {opt}")
    
    print("\n" + "="*70)
    print("STEP 3: Starting extraction...")
    print("="*70)
    
    # Run extraction
    results = extractor.run_extraction(job_family_mappings)
    
    print("\n✨ ALL DONE! Check the 'output' folder for your files.")
    
    if results:
        print("\n📁 Files created:")
        for job_family in results.keys():
            safe_name = re.sub(r'[^\w\s-]', '', job_family).replace(' ', '_')
            print(f"   - SkillBridge_{safe_name}.xlsx")
            print(f"   - SkillBridge_{safe_name}.csv")
            print(f"   - SkillBridge_{safe_name}_[timestamp].xlsx (backup)")
    else:
        print("\n❌ No data was extracted. Please check the error messages above.")

if __name__ == "__main__":
    main()
