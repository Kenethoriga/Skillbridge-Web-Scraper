"""
SkillBridge Data Extractor - Filter-Triggered Version
Interacts with dropdowns to load data, then extracts
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import pandas as pd
from datetime import datetime
import time
import os
import re
from openpyxl.styles import Font, Alignment, PatternFill

class SkillBridgeExtractor:
    def __init__(self, headless=False):
        self.headless = headless
        self.driver = None
        self.wait = None
        self.base_url = "https://skillbridge.osd.mil/locations.htm"
        self.setup_driver()
    
    def setup_driver(self):
        """Setup Chrome driver"""
        print("🛠️  Setting up Chrome driver...")
        chrome_options = Options()
        
        if self.headless:
            chrome_options.add_argument("--headless=new")
        
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            self.wait = WebDriverWait(self.driver, 60)
            print("✅ Chrome driver setup complete")
        except Exception as e:
            print(f"❌ Failed to setup Chrome driver: {e}")
            raise
    
    def navigate_to_page(self):
        """Navigate to SkillBridge locations page"""
        print(f"\n🌐 Navigating to {self.base_url}...")
        self.driver.get(self.base_url)
        time.sleep(5)  # Wait for initial load
        print("✅ Page loaded")
    
    def wait_for_page_ready(self):
        """Wait for page to be fully loaded"""
        print("⏳ Waiting for page to be ready...")
        
        # Wait for jQuery and DataTables
        self.wait.until(lambda d: d.execute_script(
            "return typeof jQuery !== 'undefined' && typeof jQuery.fn.DataTable !== 'undefined'"
        ))
        
        # Wait for the industries dropdown (Job Family)
        self.wait.until(EC.presence_of_element_located((By.ID, "industries-dropdown")))
        
        time.sleep(3)
        print("✅ Page ready")
    
    def get_industry_options(self):
        """Get all available options from industries dropdown"""
        print("\n🔍 Checking available Job Family options...")
        
        try:
            dropdown = Select(self.driver.find_element(By.ID, "industries-dropdown"))
            options = dropdown.options
            
            # Get option texts (skip first if it's "All" or empty)
            available_options = []
            for option in options:
                text = option.text.strip()
                if text and text.lower() not in ['all', 'select', '']:
                    available_options.append(text)
            
            print(f"✅ Found {len(available_options)} job family options:")
            for opt in available_options:
                print(f"   - {opt}")
            
            return available_options
            
        except Exception as e:
            print(f"❌ Error getting dropdown options: {e}")
            return []
    
    def select_industry(self, industry_name):
        """Select a specific industry from dropdown"""
        print(f"\n🎯 Selecting industry: {industry_name}")
        
        try:
            # Find the dropdown
            dropdown = Select(self.driver.find_element(By.ID, "industries-dropdown"))
            
            # Try exact match first
            try:
                dropdown.select_by_visible_text(industry_name)
                print(f"✅ Selected: {industry_name}")
                return True
            except:
                # Try partial match
                for option in dropdown.options:
                    if industry_name.lower() in option.text.lower():
                        dropdown.select_by_visible_text(option.text)
                        print(f"✅ Selected: {option.text}")
                        return True
                
                print(f"⚠️  Could not find '{industry_name}' in dropdown")
                return False
                
        except Exception as e:
            print(f"❌ Error selecting industry: {e}")
            return False
    
    def trigger_search(self):
        """Trigger the search to load data"""
        print("🔍 Triggering search...")
        
        try:
            # Look for search/submit button
            button_selectors = [
                "#search-button",
                "button[type='submit']",
                "button.search",
                "input[type='submit']",
                ".btn-search",
                "button:contains('Search')"
            ]
            
            for selector in button_selectors:
                try:
                    button = self.driver.find_element(By.CSS_SELECTOR, selector)
                    button.click()
                    print(f"✅ Clicked search button")
                    time.sleep(3)
                    return True
                except:
                    continue
            
            # If no button found, the table might auto-update on dropdown change
            print("ℹ️  No search button found - table may auto-update")
            time.sleep(5)  # Give time for auto-update
            return True
            
        except Exception as e:
            print(f"⚠️  Error triggering search: {e}")
            return False
    
    def wait_for_data_load(self):
        """Wait for data to load in table"""
        print("⏳ Waiting for data to load...")
        
        max_wait = 30
        start_time = time.time()
        
        while time.time() - start_time < max_wait:
            try:
                # Check if table has data rows
                table = self.driver.find_element(By.ID, "location-table")
                tbody = table.find_element(By.TAG_NAME, "tbody")
                rows = tbody.find_elements(By.TAG_NAME, "tr")
                
                # Filter out "no data" or loading messages
                data_rows = [r for r in rows if r.is_displayed() and 
                           'dataTables_empty' not in r.get_attribute('class')]
                
                if len(data_rows) > 0:
                    print(f"✅ Data loaded - {len(data_rows)} visible rows")
                    time.sleep(2)  # Let it fully render
                    return len(data_rows)
                
            except:
                pass
            
            time.sleep(1)
        
        print("⚠️  Timeout waiting for data")
        return 0
    
    def extract_table_data(self, job_family_name):
        """Extract all visible data from table"""
        print(f"\n📊 Extracting data for {job_family_name}...")
        
        try:
            table = self.driver.find_element(By.ID, "location-table")
            
            # Get headers - try multiple approaches
            headers = []
            try:
                # Method 1: From thead
                thead = table.find_element(By.TAG_NAME, "thead")
                header_cells = thead.find_elements(By.TAG_NAME, "th")
                headers = [h.text.strip() for h in header_cells if h.text.strip()]
            except:
                pass
            
            if not headers:
                # Method 2: Use DataTables API
                try:
                    headers_script = """
                    var table = jQuery('#location-table').DataTable();
                    var headers = [];
                    table.columns().every(function() {
                        headers.push(jQuery(this.header()).text().trim());
                    });
                    return headers;
                    """
                    headers = self.driver.execute_script(headers_script)
                    headers = [h for h in headers if h]
                except:
                    pass
            
            if not headers:
                print("⚠️  Could not extract headers, using generic names")
                headers = [f"Column_{i}" for i in range(20)]
            
            print(f"   Found {len(headers)} columns")
            
            # Extract row data using DataTables API
            print("   Extracting rows using DataTables API...")
            
            extract_script = """
            var table = jQuery('#location-table').DataTable();
            var data = [];
            
            // Get all rows (including filtered)
            table.rows({page: 'all', search: 'applied'}).every(function() {
                var rowData = this.data();
                data.push(rowData);
            });
            
            return data;
            """
            
            try:
                rows_data = self.driver.execute_script(extract_script)
                
                if rows_data and len(rows_data) > 0:
                    print(f"✅ Extracted {len(rows_data)} rows via DataTables API")
                    
                    # Convert to dict format
                    programs = []
                    for row_data in rows_data:
                        program = {}
                        for i, value in enumerate(row_data):
                            header = headers[i] if i < len(headers) else f"Column_{i}"
                            # Clean HTML tags if present
                            if isinstance(value, str):
                                value = re.sub('<[^<]+?>', '', value).strip()
                            program[header] = value
                        
                        program['Job Family'] = job_family_name
                        program['Last Updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        programs.append(program)
                    
                    return programs
                    
            except Exception as e:
                print(f"⚠️  DataTables API extraction failed: {e}")
            
            # Fallback: Extract from visible HTML
            print("   Falling back to HTML extraction...")
            return self.extract_from_html(table, headers, job_family_name)
            
        except Exception as e:
            print(f"❌ Error extracting table data: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def extract_from_html(self, table, headers, job_family_name):
        """Fallback: Extract from HTML table"""
        print("   Extracting from visible HTML rows...")
        
        programs = []
        
        try:
            tbody = table.find_element(By.TAG_NAME, "tbody")
            rows = tbody.find_elements(By.TAG_NAME, "tr")
            
            for row in rows:
                try:
                    if not row.is_displayed():
                        continue
                    
                    row_class = row.get_attribute('class') or ''
                    if 'dataTables_empty' in row_class:
                        continue
                    
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if not cells:
                        continue
                    
                    program = {}
                    for i, cell in enumerate(cells):
                        header = headers[i] if i < len(headers) else f"Column_{i}"
                        
                        # Get text content
                        text = cell.text.strip()
                        
                        # Try to extract links
                        try:
                            links = cell.find_elements(By.TAG_NAME, "a")
                            if links:
                                program[f"{header}_Link"] = links[0].get_attribute('href')
                        except:
                            pass
                        
                        program[header] = text
                    
                    program['Job Family'] = job_family_name
                    program['Last Updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    programs.append(program)
                    
                except Exception as e:
                    continue
            
            print(f"✅ Extracted {len(programs)} programs from HTML")
            return programs
            
        except Exception as e:
            print(f"❌ HTML extraction error: {e}")
            return []
    
    def extract_job_family(self, job_family_name, job_family_option):
        """Extract data for one job family"""
        print(f"\n{'='*70}")
        print(f"  EXTRACTING: {job_family_name}")
        print(f"{'='*70}")
        
        try:
            # Reset page
            self.driver.refresh()
            self.wait_for_page_ready()
            
            # Select the industry
            if not self.select_industry(job_family_option):
                return []
            
            # Trigger search
            self.trigger_search()
            
            # Wait for data to load
            row_count = self.wait_for_data_load()
            
            if row_count == 0:
                print(f"⚠️  No data loaded for {job_family_name}")
                return []
            
            # Take screenshot
            screenshot_file = f"screenshot_{job_family_name.replace(' ', '_')}.png"
            self.driver.save_screenshot(screenshot_file)
            print(f"📸 Screenshot: {screenshot_file}")
            
            # Extract data
            programs = self.extract_table_data(job_family_name)
            
            return programs
            
        except Exception as e:
            print(f"❌ Error extracting {job_family_name}: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def consolidate_locations(self, df):
        """Consolidate multiple locations for same organization/program"""
        if df.empty:
            return df
        
        print(f"\n🔄 Consolidating locations...")
        print(f"   Before: {len(df)} rows")
        
        # Find key columns (avoid duplicates)
        org_cols = []
        prog_cols = []
        loc_cols = []
        
        for col in df.columns:
            col_lower = col.lower()
            if any(term in col_lower for term in ['partner', 'organization', 'employer', 'company']) and col not in org_cols:
                org_cols.append(col)
            elif any(term in col_lower for term in ['program', 'opportunity']) and col not in prog_cols:
                prog_cols.append(col)
            elif any(term in col_lower for term in ['location', 'city', 'state', 'zip', 'address']) and col not in loc_cols:
                loc_cols.append(col)
        
        if not org_cols:
            print("⚠️  Could not identify organization column - skipping consolidation")
            return df
        
        # Build unique group columns
        group_cols = list(set(org_cols + prog_cols))
        print(f"   Grouping by: {', '.join(group_cols)}")
        
        # Add other non-location columns (avoid duplicates)
        for col in df.columns:
            if (col not in group_cols and col not in loc_cols and 
                col not in ['Last Updated', 'Job Family'] and '_Link' not in col):
                group_cols.append(col)
        
        # Remove any duplicates from group_cols
        group_cols = list(dict.fromkeys(group_cols))
        
        # Aggregation
        agg_dict = {}
        for col in df.columns:
            if col in loc_cols:
                agg_dict[col] = lambda x: ' | '.join(sorted(set(str(v).strip() 
                                                                 for v in x if pd.notna(v) and str(v).strip())))
            elif col == 'Last Updated':
                agg_dict[col] = 'max'
            elif col == 'Job Family':
                agg_dict[col] = 'first'
            elif col not in group_cols:
                agg_dict[col] = 'first'
        
        try:
            consolidated = df.groupby(group_cols, as_index=False).agg(agg_dict)
            print(f"   After: {len(consolidated)} rows")
            print(f"✅ Consolidated {len(df) - len(consolidated)} duplicates")
            return consolidated
        except Exception as e:
            print(f"⚠️  Consolidation error: {e}")
            print(f"   Skipping consolidation, returning original data")
            return df
    
    def save_to_excel(self, df, filename, sheet_name):
        """Save to Excel with formatting"""
        if df.empty:
            print(f"⚠️  No data to save")
            return
        
        print(f"\n💾 Saving to {filename}...")
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            
            worksheet = writer.sheets[sheet_name[:31]]
            
            # Header formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Auto-adjust columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = max(adjusted_width, 12)
            
            # Text wrapping
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            worksheet.freeze_panes = 'A2'
        
        print(f"✅ Saved {len(df)} programs")
    
    def run_extraction(self, job_family_mappings, output_dir='output'):
        """Main extraction process"""
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        try:
            # Navigate and prepare
            self.navigate_to_page()
            self.wait_for_page_ready()
            
            # Get available options
            available_options = self.get_industry_options()
            
            # Process each job family
            all_results = {}
            
            for job_family_name, job_family_option in job_family_mappings.items():
                programs = self.extract_job_family(job_family_name, job_family_option)
                
                if programs:
                    # Convert to DataFrame
                    df = pd.DataFrame(programs)
                    
                    # Consolidate
                    df = self.consolidate_locations(df)
                    
                    # Save
                    safe_name = re.sub(r'[^\w\s-]', '', job_family_name).replace(' ', '_')
                    
                    # Main file
                    main_file = os.path.join(output_dir, f"SkillBridge_{safe_name}.xlsx")
                    self.save_to_excel(df, main_file, job_family_name)
                    
                    # Backup
                    backup_file = os.path.join(output_dir, f"SkillBridge_{safe_name}_{timestamp}.xlsx")
                    self.save_to_excel(df, backup_file, job_family_name)
                    
                    # CSV
                    csv_file = main_file.replace('.xlsx', '.csv')
                    df.to_csv(csv_file, index=False)
                    print(f"✅ CSV: {csv_file}")
                    
                    all_results[job_family_name] = df
            
            # Summary
            print(f"\n{'='*70}")
            print("  EXTRACTION COMPLETE!")
            print(f"{'='*70}")
            for name, df in all_results.items():
                print(f"✅ {name}: {len(df)} programs")
            print(f"\n📁 Files saved to: {os.path.abspath(output_dir)}")
            print(f"{'='*70}")
            
        except Exception as e:
            print(f"\n❌ FATAL ERROR: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            if not self.headless:
                print("\n⏳ Browser will close in 10 seconds...")
                time.sleep(10)
            self.close()
    
    def close(self):
        """Close browser"""
        if self.driver:
            self.driver.quit()
            print("🔚 Browser closed")

def main():
    print("""
    ╔══════════════════════════════════════════════════════════════╗
    ║     SKILLBRIDGE DATA EXTRACTOR - DROPDOWN VERSION            ║
    ║     Extracts: Sales & Business/Financial Operations          ║
    ╚══════════════════════════════════════════════════════════════╝
    """)
    
    # Map friendly names to exact dropdown options
    # The exact text from the dropdown list
    job_family_mappings = {
        'Sales and Related': 'Sales and Related',
        'Business and Financial Operations': 'Business and Financial Operations'
    }
    
    print("\n📋 NOTE: The script will first show you ALL available job family options")
    print("   from the dropdown. If 'Business and Financial Operations' appears")
    print("   with different text, we'll update the script.\n")
    
    # User choice
    print("Run Options:")
    print("1. Headless mode (background)")
    print("2. Visible mode (RECOMMENDED - see it work!)")
    
    try:
        choice = input("\nSelect (1 or 2, default=2): ").strip() or '2'
        headless = choice == '1'
    except:
        headless = False
    
    # Run extraction
    extractor = SkillBridgeExtractor(headless=headless)
    
    print("\n" + "="*70)
    print("STEP 1: Checking available dropdown options...")
    print("="*70)
    
    extractor.navigate_to_page()
    extractor.wait_for_page_ready()
    available_options = extractor.get_industry_options()
    
    # Check if our target options exist
    print("\n" + "="*70)
    print("STEP 2: Verifying target job families...")
    print("="*70)
    
    # Find Business and Financial Operations with fuzzy matching
    business_match = None
    for opt in available_options:
        if 'business' in opt.lower() and 'financial' in opt.lower():
            business_match = opt
            print(f"✅ Found 'Business and Financial Operations' as: '{opt}'")
            job_family_mappings['Business and Financial Operations'] = opt
            break
    
    if not business_match:
        print("⚠️  'Business and Financial Operations' not found in dropdown")
        print("   Available options containing 'Business':")
        for opt in available_options:
            if 'business' in opt.lower():
                print(f"   - {opt}")
    
    # Check Sales
    sales_match = None
    for opt in available_options:
        if 'sales' in opt.lower():
            sales_match = opt
            print(f"✅ Found 'Sales and Related' as: '{opt}'")
            job_family_mappings['Sales and Related'] = opt
            break
    
    print("\n" + "="*70)
    print("STEP 3: Starting extraction...")
    print("="*70)
    
    # Run extraction
    extractor.run_extraction(job_family_mappings)
    
    print("\n✨ ALL DONE! Check the 'output' folder for your files.")
    print("\n📁 Files created:")
    print("   - SkillBridge_Sales_and_Related.xlsx")
    print("   - SkillBridge_Sales_and_Related.csv")
    if business_match:
        print("   - SkillBridge_Business_and_Financial_Operations.xlsx")
        print("   - SkillBridge_Business_and_Financial_Operations.csv")

if __name__ == "__main__":
    main()